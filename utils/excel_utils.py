import openpyxl
from openpyxl import utils
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment, Protection, Side
from openpyxl.drawing.image import Image
from logging import getLogger
my_logger = getLogger()

from utils.error_handlers import handle_errors
class ExcelUtils():
    """
    エクセル操作に関するロジック
    """
    def __init__(self, path:str):
        self.wb:Workbook = openpyxl.load_workbook(path)
        self.ws:Worksheet = self.wb.worksheets[0]

    def save_wb(self, path:str):
        '''
        指定したパスにエクセルを保存
        '''
        self.wb.save(path)
    
    def read_wooksheet(self, sheet_name:str):
        '''
        指定したタイトルのワークシートを取得
        '''
        self.ws = self.wb[sheet_name]
    
    @handle_errors
    def set_value_to_address(self, address:str, value):
        '''
        指定したセルの番地に値を書き込む
        '''
        self.ws[address] = value

    @handle_errors
    def set_money_format(self, column:int):
        '''
        指定した列のフォーマットをカンマ区切りにする
        '''
        for row in self.ws[column][1:]:  # 先頭行はヘッダーなので除外
            row.number_format = "#,##0"
        
    @handle_errors
    def merge_cells(self,range:str):
        '''
        指定した範囲のセルを結合
        '''
        self.ws.merge_cells(range)

    @handle_errors
    def unmerge_cells(self, range:str):
        '''
        指定した範囲のセルの結合を解除
        '''
        self.ws.unmerge_cells(range)

    @handle_errors
    def set_value_to_merged_cell(self, start_address:str, end_address:str, value):
        '''
        指定した結合セルに値を書き込む
        '''
        # 値を書き込むため結合解除
        range = f"{start_address}:{end_address}"
        self.unmerge_cells(range)
        self.ws[start_address] = value
        
        # 値を書き込んだ後再結合
        self.merge_cells(range)

    @handle_errors
    def set_value_alignment_to_merged_cell(self, start_address:str, end_address:str, value, horizontial="center", font="Meiryo UI", size=11, is_bold=False):
        '''
        指定した結合セルに値を書き込む
        '''
        self.set_value_to_merged_cell(start_address=start_address, end_address=end_address, value=value)
        self.ws[start_address].font = Font(name=font, size=size, bold=is_bold)
        self.ws[start_address].alignment = Alignment(horizontal=horizontial, vertical="center")
    
    @handle_errors
    def insert_rows(self, idx:int, amount:int):
        '''
        指定した位置に指定した個数の行を挿入
        '''
        self.ws.insert_rows(idx, amount)

    @handle_errors
    def set_row_height(self, idx:int, height:float):
        '''
        指定した行の高さを設定
        '''
        self.ws.row_dimensions[idx].height = height

    @handle_errors
    def get_print_area(self):
        '''
        現シートの印刷範囲取得
        '''
        return self.ws.print_area

    @handle_errors
    def set_print_area(self, range:str):
        '''
        現シートの印刷範囲設定
        '''
        self.ws.print_area = range


    # def set_header_border(self, idx:int):
    #     '''
    #     指定した行のボーダーを設定
    #     '''
    #     # self.ws.row_dimensions[idx].border = Border(top=Side(border_style="thin"))
    #     bottom_border = Border(bottom=Side(border_style="thin"))
    #     for cell in self.ws[idx]:
    #         cell.border = bottom_border
    
    @handle_errors
    def copy_row(self, src:int, dest:int):
        """
        指定された行のデータをコピーして別の行に貼り付ける関数
        :param src: int
            コピー元の行インデックス
        :param dest: int
            貼り付け先の行インデックス
        """
        for i in range(1, self.ws.max_column+1):
            cell_source = self.ws.cell(row=src, column=i)
            cell_target = self.ws.cell(row=dest, column=i)

            # スタイルを個別にコピーする
            cell_target.font = Font(
                name=cell_source.font.name,
                size=cell_source.font.size,
                bold=cell_source.font.bold,
                italic=cell_source.font.italic,
                vertAlign=cell_source.font.vertAlign,
                underline=cell_source.font.underline,
                strike=cell_source.font.strike,
                color=cell_source.font.color
            )
            cell_target.border = Border(
                left=cell_source.border.left,
                right=cell_source.border.right,
                top=cell_source.border.top,
                bottom=cell_source.border.bottom,
                diagonal=cell_source.border.diagonal,
                diagonal_direction=cell_source.border.diagonal_direction,
                outline=cell_source.border.outline,
                vertical=cell_source.border.vertical,
                horizontal=cell_source.border.horizontal
            )
            cell_target.fill = PatternFill(
                fill_type=cell_source.fill.fill_type,
                start_color=cell_source.fill.start_color,
                end_color=cell_source.fill.end_color
            )
            cell_target.number_format = cell_source.number_format
            cell_target.protection = Protection(
                locked=cell_source.protection.locked,
                hidden=cell_source.protection.hidden
            )
            cell_target.alignment = Alignment(
                horizontal=cell_source.alignment.horizontal,
                vertical=cell_source.alignment.vertical,
                text_rotation=cell_source.alignment.text_rotation,
                wrap_text=cell_source.alignment.wrap_text,
                shrink_to_fit=cell_source.alignment.shrink_to_fit,
                indent=cell_source.alignment.indent
            )
            cell_target.value = cell_source.value

    @handle_errors
    def copy_column(self, src_col, dest_col, row_index):
        """
        指定された列のデータをコピーして別の列に貼り付ける関数
        :param ws: Worksheet
            操作対象のワークシート
        :param src_col: int
            コピー元の列インデックス（1から始まる）
        :param dest_col: int
            貼り付け先の列インデックス（1から始まる）
        """
        i = row_index
        cell_source = self.ws.cell(row=i, column=src_col)
        cell_target = self.ws.cell(row=i, column=dest_col)

        # スタイルを個別にコピーする
        cell_target.font = Font(
            name=cell_source.font.name,
            size=cell_source.font.size,
            bold=cell_source.font.bold,
            italic=cell_source.font.italic,
            vertAlign=cell_source.font.vertAlign,
            underline=cell_source.font.underline,
            strike=cell_source.font.strike,
            color=cell_source.font.color
        )
        cell_target.border = Border(
            left=cell_source.border.left,
            right=cell_source.border.right,
            top=cell_source.border.top,
            bottom=cell_source.border.bottom,
            diagonal=cell_source.border.diagonal,
            diagonal_direction=cell_source.border.diagonal_direction,
            outline=cell_source.border.outline,
            vertical=cell_source.border.vertical,
            horizontal=cell_source.border.horizontal
        )
        cell_target.fill = PatternFill(
            fill_type=cell_source.fill.fill_type,
            start_color=cell_source.fill.start_color,
            end_color=cell_source.fill.end_color
        )
        cell_target.number_format = cell_source.number_format
        cell_target.protection = Protection(
            locked=cell_source.protection.locked,
            hidden=cell_source.protection.hidden
        )
        cell_target.alignment = Alignment(
            horizontal=cell_source.alignment.horizontal,
            vertical=cell_source.alignment.vertical,
            text_rotation=cell_source.alignment.text_rotation,
            wrap_text=cell_source.alignment.wrap_text,
            shrink_to_fit=cell_source.alignment.shrink_to_fit,
            indent=cell_source.alignment.indent
        )

    @handle_errors
    def delete_row(self, row:int):
        '''
        指定した行を削除
        '''
        self.ws.delete_rows(row)

    @handle_errors
    def delete_col(self, col:int):
        '''
        指定した列を削除
        '''
        self.ws.delete_cols(col)

    @handle_errors
    def set_sheet_title(self, title:str):
        '''
        シートにタイトルを設定
        '''
        self.ws.title = title

    @handle_errors
    def col_num_2_col_letter(self, col:int):
        '''
        列番号を列のアルファベット表記に変換
        '''
        return utils.cell.get_column_letter(col)
    
    @handle_errors
    def get_max_col(self):
        '''
        ワークシートの最大列数を取得する
        '''
        return self.ws.max_column
    
    @handle_errors
    def set_bgcolor_to_cell(self, row:int, col:int, bgcolor):
        '''
        指定されたセルの背景色を設定
        '''
        self.ws.cell(row=row, column=col).fill = PatternFill(patternType="solid", fgColor=bgcolor)
    
    @handle_errors
    def set_font_to_cell(self, row:int, col:int, size:int, italic:bool=False, name:str|None=None):
        '''
        指定されたセルのフォントを設定
        '''
        self.ws.cell(row=row, column=col).font = Font(name=name, size=size, italic=italic)

    @handle_errors
    def set_aliginment_center_to_cell(self, row:int, col:int):
        '''
        指定したセルを中央ぞろえにする
        '''
        self.ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

    @handle_errors
    def clear_border_to_cell(self, row:int, col:int):
        '''
        指定したセルの枠線をクリア
        '''
        self.ws.cell(row=row, column=col).border = Border(left=Side(), right=Side(), top=Side(), bottom=Side())

    @handle_errors
    def draw_diagonal(self, row:int, col:int):
        '''
        指定したセルに斜線を引く
        '''
        self.ws.cell(row=row, column=col).border = Border(right=Side(style="thin"), diagonal=Side(style="medium"), diagonalDown=True)

    @handle_errors
    def set_image_to_cell(self, address:str, path:str):
        '''
        指定したセルに画像(電子印）を書き込む
        '''
        img = Image(path)
        self.ws.add_image(img, address)